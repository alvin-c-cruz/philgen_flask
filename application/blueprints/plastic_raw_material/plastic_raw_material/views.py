from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, g, send_file, current_app
from pathlib import Path
from openpyxl import Workbook
import json
from sqlalchemy.exc import IntegrityError
from .models import PlasticRawMaterial as Obj
from .models import UserPlasticRawMaterial as Preparer
from .models import AdminPlasticRawMaterial as Approver
from .forms import Form
from application.extensions import db, month_first_day, month_last_day
from application.blueprints.user import login_required, roles_accepted
from flask_login import current_user
from .extensions import Inventory

from . import app_name, app_label


bp = Blueprint(app_name, __name__, template_folder="pages", url_prefix=f"/{app_name}")
ROLES_ACCEPTED = app_label


@bp.route("/", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def home():
    if request.method == "POST":
        date_from = request.form.get("date_from")
        date_to = request.form.get("date_to")
    else:
        date_from = month_first_day()
        date_to = month_last_day()

    records = Obj.query.order_by(getattr(Obj, f"{app_name}_name")).all()

    rows = []
    for record in records:
        record.inventory = Inventory(record, date_from=date_from, date_to=date_to)
        rows.append(record)


    context = {
        "rows": rows,
        "date_from": date_from,
        "date_to": date_to
    }

    return render_template(f"{app_name}/home.html", **context)


@bp.route("/add", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def add():
    if request.method == "POST":
        form = Form()
        form._post(request.form)
        form.user_prepare_id = current_user.id

        if form._validate_on_submit():
            form._save()
            return redirect(url_for(f'{app_name}.home'))
    else:
        form = Form()

    context = {
        "form": form,
    }

    return render_template(f"{app_name}/form.html", **context)


@bp.route(f"/edit/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def edit(record_id):   
    if request.method == "POST":
        form = Form()
        form._post(request.form)
        form.user_prepare_id = current_user.id

        if form._validate_on_submit():
            form._save()
            return redirect(url_for(f'{app_name}.home'))

    else:
        obj = Obj.query.get(record_id)
        form = Form()
        form._populate(obj)

    context = {
        "form": form,
    }

    return render_template(f"{app_name}/form.html", **context)


@bp.route("/delete/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def delete(record_id):   
    obj = Obj.query.get_or_404(record_id)
    preparer = obj.preparer
    try:
        db.session.delete(preparer)
        db.session.delete(obj)
        db.session.commit()
        flash(f"{obj} has been deleted.", category="success")
    except IntegrityError:
        db.session.rollback()
        flash(f"Cannot delete {obj} because it has related records.", category="error")

    return redirect(url_for(f'{app_name}.home'))


@bp.route("/approve/<int:record_id>", methods=['GET'])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve(record_id):
    if not current_user.admin:
        flash("Administrator rights required.", category="error")
        return redirect(url_for("plastic_raw_material.home"))
    
    obj = Obj.query.get_or_404(record_id)

    approve = Approver(
        plastic_raw_material_id=record_id,
        user_id=current_user.id
    )

    db.session.add(approve)
    db.session.commit()

    flash(f"Approved: {obj.plastic_raw_material_name}", category="success")
    return redirect(url_for(f'{app_name}.home'))   
    

@bp.route("/activate/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def activate(record_id):   
    obj = Obj.query.get_or_404(record_id)
    obj.active = True    

    db.session.commit()

    flash(f"{obj} has been activated.", category="error")

    return redirect(url_for(f'{app_name}.home'))


@bp.route("/deactivate/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def deactivate(record_id):   
    obj = Obj.query.get_or_404(record_id)
    obj.active = False    

    db.session.commit()

    flash(f"{obj} has been deactivated.", category="error")

    return redirect(url_for(f'{app_name}.home'))


@bp.route("/unlock/<int:record_id>", methods=['GET'])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def unlock(record_id):
    if not current_user.admin:
        flash("Administrator rights required.", category="error")
        redirect(url_for(f'{app_name}.home'))
    
    obj = Obj.query.get_or_404(record_id)

    approve = Approver.query.filter_by(plastic_raw_material_id=record_id).first()
    
    db.session.delete(approve)
    db.session.commit()

    flash(f"Unlocked: {getattr(obj,f'{app_name}_name')}", category="error")
    return redirect(url_for(f'{app_name}.home'))      


@bp.route("/_autocomplete", methods=['GET'])
def autocomplete():
    rows = [row for row in Obj.query.order_by(Obj.plastic_raw_material_name).all()]
    return Response(json.dumps(rows), mimetype='application/json')




@bp.route("/download")
@login_required
@roles_accepted([ROLES_ACCEPTED])
def download():
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    data = Obj.query.all()

    filename = create_file(current_app=current_app, data=data, date_from=date_from, date_to=date_to)
    return send_file('{}'.format(filename), as_attachment=True)


def create_file(current_app, data, date_from, date_to):
    wb = Workbook()
    ws = wb.active
    ws.title = "SUMMARY"

    row_num = create_header(ws, current_app, data, date_from, date_to)
    row_num = create_detail(ws, current_app, data, date_from, date_to, row_num)


    create_stock_card(wb, current_app, data, date_from, date_to)

    filename = Path(current_app.instance_path) / "temp" / "plastic_raw_material.xlsx"
    wb.save(filename)
    wb.close()

    return filename


def create_header(ws, current_app, data, date_from, date_to):
    cell = ws["A1"]
    cell.value = "Rowell Industrial Corporation"

    cell = ws["A2"]
    cell.value = "Plastic Raw Material Inventory Report"

    cell = ws["A3"]
    cell.value = f"From {date_from} to {date_to}"

    columns = {
        "A": "Name",
        "B": "Code",
        "C": ""
    }

    row_num = 5
    for column, value in columns.items():
        cell = ws[f"{column}{row_num}"]
        cell.value = value

    row_num += 1

    return row_num


def create_detail(ws, current_app, data, date_from, date_to, row_num):
    for record in data:
        columns = {
            "A": record.plastic_raw_material_name,
            "B": record.plastic_raw_material_code,
            "C": "Go to"
        }

        for column, value in columns.items():
            cell = ws[f"{column}{row_num}"]
            cell.value = value

            if column == "C":
                cell.hyperlink = f"#'{record.plastic_raw_material_code}'!A1"
                cell.style = "Hyperlink"

        row_num += 1
    
    return row_num



def create_stock_card(wb, current_app, data, date_from, date_to):
    for record in data:
        if not record.active: continue
        sheet_name = record.plastic_raw_material_code
        wb.create_sheet(title=sheet_name)
        ws = wb[sheet_name]

        row_num = create_stock_card_header(ws, current_app, record, date_from, date_to)
        row_num = create_stock_card_detail(ws, current_app, record, date_from, date_to, row_num)


def create_stock_card_header(ws, current_app, record, date_from, date_to):
    cell = ws["E1"]
    cell.value = "Back to Summary"
    cell.hyperlink = f"#'SUMMARY'!A1"
    cell.style = "Hyperlink"

    cell = ws["A1"]
    cell.value = "Rowell Industrial Corporation"

    cell = ws["A2"]
    cell.value = "Plastic Raw Material Stock Card"

    cell = ws["A3"]
    cell.value = f"From {date_from} to {date_to}"

    cell = ws["A5"]
    cell.value = "Plastic Raw Material Name"

    cell = ws["B5"]
    cell.value = record.plastic_raw_material_name

    cell = ws["A6"]
    cell.value = "Plastic Raw Material Code"

    cell = ws["B6"]
    cell.value = record.plastic_raw_material_code

    # Detail headers
    headers = {
        "A": "Date",
        # "B": "Beginning",
        # "C": "Received",
        "B": "Issued",
        # "E": "Returned",
        # "F": "Defective",
        # "G": "Adjustment",
        # "H": "Ending",
    }

    row_num = 8
    for column, value in headers.items():
        cell = ws[f"{column}{row_num}"]
        cell.value = value

    row_num += 1

    return row_num


def create_stock_card_detail(ws, current_app, record, date_from, date_to, row_num):
    for issuance in record.plastic_raw_material_issuance_details:
        details = {
            "A": issuance.plastic_raw_material_issuance.record_date,
            "B": issuance.quantity,
        }

        for column, value in details.items():
            cell = ws[f"{column}{row_num}"]
            cell.value = value

        row_num += 1
